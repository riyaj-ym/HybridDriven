import openpyxl
from openpyxl.styles import PatternFill


class XLUtils:

    def __init__(self, file, sheet):
        self.file = file
        self.sheet = sheet

    def getRows(self):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        rows = ws.max_row
        return rows

    def getColumns(self):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        columns = ws.max_column
        return columns

    def readData(self, r, c):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        data = ws.cell(r, c).value
        return data

    def writeData(self, r, c, data):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        ws.cell(r, c).value = data
        wb.save(self.file)

    def greenFill(self, r, c):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        greenFill = PatternFill(start_color='94d454',
                                end_color='94d454',
                                fill_type='solid')
        ws.cell(r, c).fill = greenFill
        wb.save(self.file)

    def redFill(self, r, c):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet]
        redFill = PatternFill(start_color='fc6404',
                              end_color='fc6404',
                              fill_type='solid')
        ws.cell(r, c).fill = redFill
        wb.save(self.file)
